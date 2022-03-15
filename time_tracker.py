from sys import maxsize
import pandas
from openpyxl import Workbook, load_workbook
from datetime import datetime

workbook = load_workbook('study_time_tracker.xlsx')

sheet = workbook.active



#df = pandas.read_excel("study_time_tracker.xlsx", engine='openpyxl')
# printing all columns of the dataframe
#print(df.columns.tolist())



#Date Column
def date_time():
    for col in sheet.iter_rows(min_col=1, min_row=2, max_col=1, max_row=1000):
        for cell in col:
            print(cell, cell.value)
        
            if cell.value == None:
                current_day = datetime.today()
                cell.value = current_day
                return cell.value
                break
            

            else:
                break
    

print()
date_time()


#startTime column
def start_time():
    for col in sheet.iter_rows(min_col=2, min_row=2, max_col=2, max_row=1000):
        for cell in col:
            print(cell, cell.value)

            if cell.value == None:
                t = datetime.now() #Converts current time
                starttime = t.strftime("%H:%M")
                cell.value = starttime
                return cell.value

                
            else:
                break


print()
start_time()

print()

print("Please enter exit!\n")
exitingInput = input()
print()




while(True):
    
    #endTime column
    if (exitingInput == "exit"):
        def end_time():
            for col in sheet.iter_rows(min_col=3, min_row=2,  max_col=3, max_row=1000):
                for cell in col:
                    print(cell, cell.value)  
                    if cell.value == None: 
                        t = datetime.now()
                        endTime = t.strftime("%H:%M")
                        cell.value = endTime
                        return cell.value
                        
                    else:
                        break
        end_time()
            
        print()
        
        
        df = pandas.read_excel("study_time_tracker.xlsx", engine='openpyxl')
        #df.index = pandas.RangeIndex(start=0, stop=20, step=1)
        print(df.info())
        
        #df['startTime'] = pandas.to_datetime(df['startTime'] + ',03,2022', format="%H,%d,%m,%Y")
        df['startTime'] = pandas.to_datetime(df['startTime'] + df['date'].dt.strftime(",%d,%m,%Y"), format='%H:%M,%d,%m,%Y')
        df['endTime'] = pandas.to_datetime(df['endTime'] + df['date'].dt.strftime(",%d,%m,%Y"), format='%H:%M,%d,%m,%Y')
        df["totalTime"] = df['endTime'] - df["startTime"]
        print()
        print(df)
        print()
        df.to_excel("new_tracker.xlsx")
                    
        
        #print(df.info)
        
    

        '''
        def total_time():
            while True:
                df = pandas.read_excel("study_time_tracker.xlsx", engine='openpyxl')
        #df["totalTime"] = int(df["endTime"]) - int(df["startTime"]
                
                
                if df[df["startTime"]] == None | df[df["startTime"]] == "NA" | df[df["startTime"]] == "NaN":
                    break
                elif(df[df["endTime"]] == None | df[df["endTime"]] == "NA" | df[df["endTime"]] == "NaN"):
                    break
                else:
                

                start = pandas.to_datetime(df["startTime"].astype(str), format='%H:%M')
                end = pandas.to_datetime(df["endTime"].astype(str), format='%H:%M')
                df['totalTime'] = end - start
                break
            
            print()
            print(df)
        total_time()
        '''

        #df["totalTime"] = df["endTime"].subtract(df["startTime"])
        #df["totalTime"] = (pandas.to_datetime(df["endTime"], format = ("%H,%M")).subtract(pandas.to_datetime(df["startTime"], format = ("%H,%M"))))
        #print(df)

        '''
        def total_time():
            for col in sheet.iter_rows(min_col=4, min_row=2, max_col=4):
                for cell in col:
                    #if cell.value == None:
                    df = pandas.read_excel("study_time_tracker.xlsx", engine='openpyxl')
                    #df["totalTime"] = (pandas.to_datetime(df["endTime"]).subtract(pandas.to_datetime(df["startTime"]))).abs

                    print(cell, cell.value)
                        #cell.value = df["totalTime"]
                    #cell.value = df["totalTime"]
                    #else:
                     #   break
        
        total_time()
        '''

        '''
        for col in sheet.iter_rows(min_col=4, min_row=2, max_col=4):
            for cell in col:
                print(cell, cell.value)
                if cell.value == None:
                  
                   #sheet["D"] = "=SUBTRACT(C:B)"
                   #cell.value = sheet["D"]
                   break
                else:
                    break  
        
        '''
        
    
    break



workbook.save('study_time_tracker.xlsx')